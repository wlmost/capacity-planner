"""
Unit Tests für CapacityWidget Export-Funktionalität
"""
import pytest
import os
import tempfile
from datetime import datetime
from unittest.mock import Mock, MagicMock, patch
from PySide6.QtWidgets import QApplication
from PySide6.QtCore import QDate

from src.views.capacity_widget import CapacityWidget
from src.viewmodels.capacity_viewmodel import CapacityViewModel
from src.models.capacity import Capacity
from src.models.worker import Worker


@pytest.fixture
def mock_viewmodel():
    """Mock CapacityViewModel"""
    viewmodel = Mock(spec=CapacityViewModel)
    viewmodel.get_active_workers = Mock(return_value=[])
    viewmodel.load_all_capacities = Mock(return_value=[])
    
    # Mock signals
    viewmodel.capacity_created = Mock()
    viewmodel.capacity_created.connect = Mock()
    viewmodel.capacity_updated = Mock()
    viewmodel.capacity_updated.connect = Mock()
    viewmodel.capacity_deleted = Mock()
    viewmodel.capacity_deleted.connect = Mock()
    viewmodel.capacities_loaded = Mock()
    viewmodel.capacities_loaded.connect = Mock()
    viewmodel.utilization_calculated = Mock()
    viewmodel.utilization_calculated.connect = Mock()
    viewmodel.validation_failed = Mock()
    viewmodel.validation_failed.connect = Mock()
    viewmodel.error_occurred = Mock()
    viewmodel.error_occurred.connect = Mock()
    
    return viewmodel


@pytest.fixture
def sample_workers():
    """Sample Workers für Tests"""
    return [
        Worker(id=1, name="Alice", email="alice@test.com", team="Team A", active=True),
        Worker(id=2, name="Bob", email="bob@test.com", team="Team B", active=True),
    ]


@pytest.fixture
def sample_capacities():
    """Sample Capacities für Tests"""
    return [
        Capacity(
            id=1,
            worker_id=1,
            start_date=datetime(2025, 1, 1),
            end_date=datetime(2025, 1, 31),
            planned_hours=160.0,
            notes="Januar Planung"
        ),
        Capacity(
            id=2,
            worker_id=2,
            start_date=datetime(2025, 2, 1),
            end_date=datetime(2025, 2, 28),
            planned_hours=150.0,
            notes="Februar Planung"
        ),
    ]


@pytest.fixture
def capacity_widget(qtbot, mock_viewmodel):
    """CapacityWidget Fixture"""
    widget = CapacityWidget(mock_viewmodel)
    qtbot.addWidget(widget)
    return widget


class TestCapacityWidgetExportButtons:
    """Tests für Export-Buttons"""
    
    def test_export_csv_button_exists(self, capacity_widget):
        """Test: CSV Export Button existiert"""
        assert capacity_widget._export_csv_button is not None
        assert capacity_widget._export_csv_button.text() == "📊 Export CSV"
    
    def test_export_excel_button_exists(self, capacity_widget):
        """Test: Excel Export Button existiert"""
        assert capacity_widget._export_excel_button is not None
        assert capacity_widget._export_excel_button.text() == "📗 Export Excel"


class TestCapacityWidgetCSVExport:
    """Tests für CSV Export"""
    
    def test_export_csv_no_data_shows_warning(self, capacity_widget, qtbot):
        """Test: Warnung bei fehlenden Daten"""
        capacity_widget._capacities = []
        
        with patch('PySide6.QtWidgets.QMessageBox.warning') as mock_warning:
            capacity_widget._export_to_csv()
            mock_warning.assert_called_once()
            args = mock_warning.call_args[0]
            assert "Keine Daten" in args[1]
    
    def test_export_csv_with_data(self, capacity_widget, sample_workers, sample_capacities, qtbot):
        """Test: CSV Export mit Daten"""
        capacity_widget._capacities = sample_capacities
        capacity_widget._workers = sample_workers
        
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv') as tmp_file:
            temp_path = tmp_file.name
        
        try:
            with patch('PySide6.QtWidgets.QFileDialog.getSaveFileName', return_value=(temp_path, '')):
                with patch('PySide6.QtWidgets.QMessageBox.information') as mock_info:
                    capacity_widget._export_to_csv()
                    mock_info.assert_called_once()
            
            # Verify file was created and contains data
            assert os.path.exists(temp_path)
            with open(temp_path, 'r', encoding='utf-8') as f:
                content = f.read()
                assert 'Kapazitätsplanung Bericht' in content
                assert 'Alice' in content
                assert 'Bob' in content
                assert '160.0' in content
                assert '150.0' in content
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
    
    def test_export_csv_cancelled(self, capacity_widget, sample_capacities, qtbot):
        """Test: CSV Export abgebrochen"""
        capacity_widget._capacities = sample_capacities
        
        with patch('PySide6.QtWidgets.QFileDialog.getSaveFileName', return_value=('', '')):
            with patch('PySide6.QtWidgets.QMessageBox.information') as mock_info:
                capacity_widget._export_to_csv()
                # Should not show info message when cancelled
                mock_info.assert_not_called()


class TestCapacityWidgetExcelExport:
    """Tests für Excel Export"""
    
    def test_export_excel_no_data_shows_warning(self, capacity_widget, qtbot):
        """Test: Warnung bei fehlenden Daten"""
        capacity_widget._capacities = []
        
        with patch('PySide6.QtWidgets.QMessageBox.warning') as mock_warning:
            capacity_widget._export_to_excel()
            mock_warning.assert_called_once()
            args = mock_warning.call_args[0]
            assert "Keine Daten" in args[1]
    
    def test_export_excel_with_data(self, capacity_widget, sample_workers, sample_capacities, qtbot):
        """Test: Excel Export mit Daten"""
        capacity_widget._capacities = sample_capacities
        capacity_widget._workers = sample_workers
        
        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name
        
        try:
            with patch('PySide6.QtWidgets.QFileDialog.getSaveFileName', return_value=(temp_path, '')):
                with patch('PySide6.QtWidgets.QMessageBox.information') as mock_info:
                    capacity_widget._export_to_excel()
                    mock_info.assert_called_once()
            
            # Verify file was created
            assert os.path.exists(temp_path)
            
            # Verify Excel file content
            from openpyxl import load_workbook
            wb = load_workbook(temp_path)
            ws = wb.active
            
            # Check title
            assert ws.cell(row=1, column=1).value == "Kapazitätsplanung Bericht"
            
            # Check headers
            assert ws.cell(row=6, column=1).value == "ID"
            assert ws.cell(row=6, column=2).value == "Worker"
            
            # Check data exists
            assert ws.cell(row=7, column=1).value == 1  # First capacity ID
            assert "Alice" in str(ws.cell(row=7, column=2).value)
            
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
    
    def test_export_excel_cancelled(self, capacity_widget, sample_capacities, qtbot):
        """Test: Excel Export abgebrochen"""
        capacity_widget._capacities = sample_capacities
        
        with patch('PySide6.QtWidgets.QFileDialog.getSaveFileName', return_value=('', '')):
            with patch('PySide6.QtWidgets.QMessageBox.information') as mock_info:
                capacity_widget._export_to_excel()
                # Should not show info message when cancelled
                mock_info.assert_not_called()
