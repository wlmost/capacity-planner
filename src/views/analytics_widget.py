"""
AnalyticsWidget - Dashboard für Auslastungs-Analysen
"""
from typing import List, Dict, Optional
from datetime import datetime, timedelta
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QPushButton,
    QDateEdit, QHeaderView, QGroupBox, QFormLayout,
    QMessageBox, QFileDialog, QProgressBar, QTabWidget,
    QComboBox
)
from PySide6.QtCore import Qt, QDate, Signal
from PySide6.QtGui import QFont, QColor

from ..services.analytics_service import AnalyticsService
from ..repositories.worker_repository import WorkerRepository
from ..repositories.time_entry_repository import TimeEntryRepository
from ..repositories.capacity_repository import CapacityRepository
from ..models.worker import Worker
from .utilization_chart_widget import UtilizationChartWidget
from .worker_detail_dialog import WorkerDetailDialog


class AnalyticsWidget(QWidget):
    """
    Widget für Analytics Dashboard
    
    Features:
        - Team-Übersicht mit Auslastung
        - Zeitraum-basierte Filterung
        - Farbkodierung für Auslastungsstatus
        - Export-Funktionalität (CSV)
    """
    
    data_refreshed = Signal()
    
    def __init__(
        self,
        analytics_service: AnalyticsService,
        worker_repository: WorkerRepository,
        time_entry_repository: TimeEntryRepository,
        capacity_repository: CapacityRepository
    ):
        super().__init__()
        self._analytics_service = analytics_service
        self._worker_repository = worker_repository
        self._time_entry_repository = time_entry_repository
        self._capacity_repository = capacity_repository
        self._workers: List[Worker] = []
        self._utilization_data: Dict[int, Dict] = {}
        
        self._setup_ui()
        self._load_initial_data()
    
    def _setup_ui(self):
        """Erstellt das UI Layout"""
        layout = QVBoxLayout(self)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("Analytics Dashboard")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Refresh Button
        self._refresh_button = QPushButton("🔄 Aktualisieren")
        self._refresh_button.clicked.connect(self._refresh_data)
        header_layout.addWidget(self._refresh_button)
        
        # Export CSV Button
        self._export_csv_button = QPushButton("📊 Export CSV")
        self._export_csv_button.clicked.connect(self._export_to_csv)
        header_layout.addWidget(self._export_csv_button)
        
        # Export Excel Button
        self._export_excel_button = QPushButton("📗 Export Excel")
        self._export_excel_button.clicked.connect(self._export_to_excel)
        header_layout.addWidget(self._export_excel_button)
        
        layout.addLayout(header_layout)
        
        # Filter Section
        filter_group = QGroupBox("Zeitraum")
        filter_layout = QHBoxLayout(filter_group)
        
        filter_layout.addWidget(QLabel("Von:"))
        self._start_date_filter = QDateEdit()
        self._start_date_filter.setCalendarPopup(True)
        self._start_date_filter.setDate(QDate.currentDate().addMonths(-1))
        self._start_date_filter.setDisplayFormat("dd.MM.yyyy")
        self._start_date_filter.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._start_date_filter)
        
        filter_layout.addWidget(QLabel("Bis:"))
        self._end_date_filter = QDateEdit()
        self._end_date_filter.setCalendarPopup(True)
        self._end_date_filter.setDate(QDate.currentDate())
        self._end_date_filter.setDisplayFormat("dd.MM.yyyy")
        self._end_date_filter.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._end_date_filter)
        
        filter_layout.addSpacing(20)
        
        # Team-Filter
        filter_layout.addWidget(QLabel("Team:"))
        self._team_filter = QComboBox()
        self._team_filter.addItem("Alle Teams", None)
        self._team_filter.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._team_filter)
        
        filter_layout.addSpacing(20)
        
        # Status-Filter
        filter_layout.addWidget(QLabel("Status:"))
        self._status_filter = QComboBox()
        self._status_filter.addItem("Alle Status", None)
        self._status_filter.addItem("⚠ Unter (< 80%)", "under")
        self._status_filter.addItem("✓ Optimal (80-110%)", "optimal")
        self._status_filter.addItem("❗ Über (> 110%)", "over")
        self._status_filter.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._status_filter)
        
        filter_layout.addStretch()
        
        layout.addWidget(filter_group)
        
        # Statistics Summary
        stats_group = self._create_statistics_group()
        layout.addWidget(stats_group)
        
        # Tab Widget für Tabelle und Chart
        self._content_tabs = QTabWidget()
        
        # Tab 1: Team Overview Table
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        
        self._team_table = QTableWidget()
        self._team_table.setColumnCount(7)
        self._team_table.setHorizontalHeaderLabels([
            "Worker", "Team", "Geplant (h)", "Gearbeitet (h)", 
            "Differenz (h)", "Auslastung (%)", "Status"
        ])
        self._team_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._team_table.setSelectionBehavior(QTableWidget.SelectRows)
        self._team_table.setAlternatingRowColors(True)
        self._team_table.setSortingEnabled(True)
        self._team_table.cellDoubleClicked.connect(self._on_worker_double_clicked)
        table_layout.addWidget(self._team_table)
        
        self._content_tabs.addTab(table_widget, "📊 Tabelle")
        
        # Tab 2: Chart
        self._chart_widget = UtilizationChartWidget()
        self._content_tabs.addTab(self._chart_widget, "📈 Diagramm")
        
        layout.addWidget(self._content_tabs)
        
        # Status Label
        self._status_label = QLabel()
        self._status_label.setWordWrap(True)
        layout.addWidget(self._status_label)
    
    def _create_statistics_group(self) -> QWidget:
        """Erstellt die Statistik-Übersicht"""
        group = QGroupBox("Zusammenfassung")
        layout = QFormLayout(group)
        
        # Total Workers
        self._total_workers_label = QLabel("0")
        self._total_workers_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addRow("Aktive Workers:", self._total_workers_label)
        
        # Total Planned Hours
        self._total_planned_label = QLabel("0.0 h")
        self._total_planned_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addRow("Gesamt Geplant:", self._total_planned_label)
        
        # Total Worked Hours
        self._total_worked_label = QLabel("0.0 h")
        self._total_worked_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addRow("Gesamt Gearbeitet:", self._total_worked_label)
        
        # Average Utilization
        self._avg_utilization_label = QLabel("0.0%")
        self._avg_utilization_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addRow("Ø Auslastung:", self._avg_utilization_label)
        
        # Progress Bar for Average
        self._avg_progress = QProgressBar()
        self._avg_progress.setMinimum(0)
        self._avg_progress.setMaximum(150)
        self._avg_progress.setTextVisible(True)
        layout.addRow("", self._avg_progress)
        
        return group
    
    def _load_initial_data(self):
        """Lädt initiale Daten"""
        try:
            self._workers = self._worker_repository.find_all()
            self._workers = [w for w in self._workers if w.active]
            
            # Team-Filter populieren
            teams = sorted(set(w.team for w in self._workers if w.team))
            for team in teams:
                self._team_filter.addItem(team, team)
            
            self._refresh_data()
        except Exception as e:
            self._show_error(f"Fehler beim Laden der Daten: {str(e)}")
    
    def _refresh_data(self):
        """Aktualisiert alle Daten"""
        self._status_label.setText("Daten werden geladen...")
        self._status_label.setStyleSheet("color: blue;")
        
        try:
            # Zeitraum ermitteln
            start_date = self._start_date_filter.date().toPython()
            end_date = self._end_date_filter.date().toPython()
            
            start_datetime = datetime(start_date.year, start_date.month, start_date.day)
            end_datetime = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
            
            # Filter anwenden
            filtered_workers = self._apply_filters()
            
            # Auslastung für jeden Worker berechnen
            self._utilization_data = {}
            for worker in filtered_workers:
                utilization = self._analytics_service.calculate_worker_utilization(
                    worker.id, start_datetime, end_datetime
                )
                if utilization:
                    self._utilization_data[worker.id] = utilization
            
            # UI aktualisieren
            self._update_statistics()
            self._update_table()
            
            self._show_success("Daten erfolgreich geladen")
            self.data_refreshed.emit()
            
        except Exception as e:
            self._show_error(f"Fehler beim Aktualisieren: {str(e)}")
    
    def _update_statistics(self):
        """Aktualisiert Statistik-Übersicht"""
        if not self._utilization_data:
            return
        
        total_workers = len(self._workers)
        total_planned = sum(data['hours_planned'] for data in self._utilization_data.values())
        total_worked = sum(data['hours_worked'] for data in self._utilization_data.values())
        
        # Durchschnittliche Auslastung berechnen
        if self._utilization_data:
            avg_utilization = sum(
                data['utilization_percent'] 
                for data in self._utilization_data.values()
            ) / len(self._utilization_data)
        else:
            avg_utilization = 0.0
        
        # Labels aktualisieren
        self._total_workers_label.setText(str(total_workers))
        self._total_planned_label.setText(f"{total_planned:.1f} h")
        self._total_worked_label.setText(f"{total_worked:.1f} h")
        self._avg_utilization_label.setText(f"{avg_utilization:.1f}%")
        
        # Progress Bar
        self._avg_progress.setValue(int(avg_utilization))
        self._avg_progress.setFormat(f"{avg_utilization:.1f}%")
        
        # Farbkodierung
        if avg_utilization < 80:
            self._avg_progress.setStyleSheet("QProgressBar::chunk { background-color: orange; }")
            self._avg_utilization_label.setStyleSheet("font-weight: bold; font-size: 14px; color: orange;")
        elif avg_utilization <= 110:
            self._avg_progress.setStyleSheet("QProgressBar::chunk { background-color: green; }")
            self._avg_utilization_label.setStyleSheet("font-weight: bold; font-size: 14px; color: green;")
        else:
            self._avg_progress.setStyleSheet("QProgressBar::chunk { background-color: red; }")
            self._avg_utilization_label.setStyleSheet("font-weight: bold; font-size: 14px; color: red;")
    
    def _update_table(self):
        """Aktualisiert Team-Tabelle und Chart"""
        self._team_table.setRowCount(0)
        
        for worker in self._workers:
            if worker.id not in self._utilization_data:
                continue
            
            data = self._utilization_data[worker.id]
            row = self._team_table.rowCount()
            self._team_table.insertRow(row)
            
            # Worker Name
            name_item = QTableWidgetItem(worker.name)
            self._team_table.setItem(row, 0, name_item)
            
            # Team
            team_item = QTableWidgetItem(worker.team or "-")
            self._team_table.setItem(row, 1, team_item)
            
            # Geplante Stunden
            planned_item = QTableWidgetItem(f"{data['hours_planned']:.1f}")
            planned_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self._team_table.setItem(row, 2, planned_item)
            
            # Gearbeitete Stunden
            worked_item = QTableWidgetItem(f"{data['hours_worked']:.1f}")
            worked_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self._team_table.setItem(row, 3, worked_item)
            
            # Differenz
            diff = data['hours_worked'] - data['hours_planned']
            diff_item = QTableWidgetItem(f"{diff:+.1f}")
            diff_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if diff < 0:
                diff_item.setForeground(QColor("orange"))
            elif diff > 0:
                diff_item.setForeground(QColor("blue"))
            self._team_table.setItem(row, 4, diff_item)
            
            # Auslastung Prozent
            util_percent = data['utilization_percent']
            util_item = QTableWidgetItem(f"{util_percent:.1f}%")
            util_item.setTextAlignment(Qt.AlignCenter)
            self._team_table.setItem(row, 5, util_item)
            
            # Status
            status_item = self._get_status_item(util_percent)
            self._team_table.setItem(row, 6, status_item)
        
        # Chart aktualisieren
        self._chart_widget.update_chart(self._workers, self._utilization_data)
    
    def _apply_filters(self) -> List[Worker]:
        """Wendet aktuelle Filter auf Worker-Liste an"""
        filtered = list(self._workers)
        
        # Team-Filter
        team_filter = self._team_filter.currentData()
        if team_filter:
            filtered = [w for w in filtered if w.team == team_filter]
        
        # Status-Filter (benötigt Utilization-Daten)
        status_filter = self._status_filter.currentData()
        if status_filter:
            # Temporär alle Utilization-Daten berechnen für Filterung
            temp_data = {}
            start_date = self._start_date_filter.date().toPython()
            end_date = self._end_date_filter.date().toPython()
            start_datetime = datetime(start_date.year, start_date.month, start_date.day)
            end_datetime = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
            
            for worker in filtered:
                utilization = self._analytics_service.calculate_worker_utilization(
                    worker.id, start_datetime, end_datetime
                )
                if utilization:
                    temp_data[worker.id] = utilization['utilization_percent']
            
            # Nach Status filtern
            if status_filter == "under":
                filtered = [w for w in filtered if temp_data.get(w.id, 0) < 80]
            elif status_filter == "optimal":
                filtered = [w for w in filtered if 80 <= temp_data.get(w.id, 0) <= 110]
            elif status_filter == "over":
                filtered = [w for w in filtered if temp_data.get(w.id, 0) > 110]
        
        return filtered
    
    def _get_status_item(self, utilization: float) -> QTableWidgetItem:
        """Erstellt Status-Item mit Farbkodierung"""
        if utilization < 80:
            item = QTableWidgetItem("⚠️ Unter")
            item.setForeground(QColor("orange"))
            item.setBackground(QColor(255, 165, 0, 50))
        elif utilization <= 110:
            item = QTableWidgetItem("✓ Optimal")
            item.setForeground(QColor("green"))
            item.setBackground(QColor(0, 255, 0, 50))
        else:
            item = QTableWidgetItem("❗ Über")
            item.setForeground(QColor("red"))
            item.setBackground(QColor(255, 0, 0, 50))
        
        item.setTextAlignment(Qt.AlignCenter)
        return item
    
    def _on_filter_changed(self):
        """Handler für Filter-Änderungen"""
        self._refresh_data()
    
    def _export_to_csv(self):
        """Exportiert Daten als CSV"""
        if not self._utilization_data:
            QMessageBox.warning(
                self, 
                "Keine Daten", 
                "Es sind keine Daten zum Exportieren vorhanden."
            )
            return
        
        # Datei-Dialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export als CSV",
            f"analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                
                # Header
                writer.writerow([
                    'Worker', 'Team', 'Geplant (h)', 'Gearbeitet (h)',
                    'Differenz (h)', 'Auslastung (%)', 'Status'
                ])
                
                # Daten
                for worker in self._workers:
                    if worker.id not in self._utilization_data:
                        continue
                    
                    data = self._utilization_data[worker.id]
                    diff = data['hours_worked'] - data['hours_planned']
                    util = data['utilization_percent']
                    
                    if util < 80:
                        status = "Unter"
                    elif util <= 110:
                        status = "Optimal"
                    else:
                        status = "Über"
                    
                    writer.writerow([
                        worker.name,
                        worker.team or "-",
                        f"{data['hours_planned']:.1f}",
                        f"{data['hours_worked']:.1f}",
                        f"{diff:+.1f}",
                        f"{util:.1f}",
                        status
                    ])
                
                # Zusammenfassung
                writer.writerow([])
                writer.writerow(['Zusammenfassung'])
                writer.writerow(['Aktive Workers', len(self._workers)])
                
                total_planned = sum(d['hours_planned'] for d in self._utilization_data.values())
                total_worked = sum(d['hours_worked'] for d in self._utilization_data.values())
                avg_util = sum(d['utilization_percent'] for d in self._utilization_data.values()) / len(self._utilization_data)
                
                writer.writerow(['Gesamt Geplant (h)', f"{total_planned:.1f}"])
                writer.writerow(['Gesamt Gearbeitet (h)', f"{total_worked:.1f}"])
                writer.writerow(['Ø Auslastung (%)', f"{avg_util:.1f}"])
            
            self._show_success(f"Export erfolgreich: {file_path}")
            
        except Exception as e:
            self._show_error(f"Fehler beim Exportieren: {str(e)}")
    
    def _export_to_excel(self):
        """Exportiert Daten als Excel mit Formatierung"""
        if not self._utilization_data:
            QMessageBox.warning(
                self, 
                "Keine Daten", 
                "Es sind keine Daten zum Exportieren vorhanden."
            )
            return
        
        # Datei-Dialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export als Excel",
            f"analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analytics"
            
            # Header-Zeile mit Formatierung
            headers = ['Worker', 'Team', 'Geplant (h)', 'Gearbeitet (h)', 
                      'Differenz (h)', 'Auslastung (%)', 'Status']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Daten-Zeilen
            row_num = 2
            for worker in self._workers:
                if worker.id not in self._utilization_data:
                    continue
                
                data = self._utilization_data[worker.id]
                diff = data['hours_worked'] - data['hours_planned']
                util = data['utilization_percent']
                
                if util < 80:
                    status = "⚠️ Unter"
                    status_color = "FFA500"  # Orange
                elif util <= 110:
                    status = "✓ Optimal"
                    status_color = "32CD32"  # Green
                else:
                    status = "❗ Über"
                    status_color = "FF4500"  # Red
                
                row_data = [
                    worker.name,
                    worker.team or "-",
                    data['hours_planned'],
                    data['hours_worked'],
                    diff,
                    util,
                    status
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center' if col_num > 2 else 'left')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Status-Spalte farblich hervorheben
                    if col_num == 7:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                    
                    # Differenz-Spalte farblich hervorheben
                    if col_num == 5:
                        if diff < 0:
                            cell.font = Font(color="FFA500")  # Orange
                        elif diff > 0:
                            cell.font = Font(color="0000FF")  # Blue
                
                row_num += 1
            
            # Zusammenfassung (mit Abstand)
            row_num += 2
            summary_row = row_num
            
            summary_cell = ws.cell(row=summary_row, column=1, value="Zusammenfassung")
            summary_cell.font = Font(bold=True, size=14, color="FFFFFF")
            summary_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
            
            total_planned = sum(d['hours_planned'] for d in self._utilization_data.values())
            total_worked = sum(d['hours_worked'] for d in self._utilization_data.values())
            avg_util = sum(d['utilization_percent'] for d in self._utilization_data.values()) / len(self._utilization_data)
            
            summary_data = [
                ('Aktive Workers', len(self._workers)),
                ('Gesamt Geplant (h)', f"{total_planned:.1f}"),
                ('Gesamt Gearbeitet (h)', f"{total_worked:.1f}"),
                ('Ø Auslastung (%)', f"{avg_util:.1f}")
            ]
            
            for i, (label, value) in enumerate(summary_data, 1):
                label_cell = ws.cell(row=summary_row + i, column=1, value=label)
                label_cell.font = Font(bold=True)
                
                value_cell = ws.cell(row=summary_row + i, column=2, value=value)
                value_cell.alignment = Alignment(horizontal='right')
                value_cell.font = Font(bold=True)
            
            # Spaltenbreiten anpassen
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            
            # Speichern
            wb.save(file_path)
            
            self._show_success(f"Excel-Export erfolgreich: {file_path}")
            
        except ImportError:
            self._show_error("openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
        except Exception as e:
            self._show_error(f"Fehler beim Excel-Export: {str(e)}")
    
    def _show_success(self, message: str):
        """Zeigt Erfolgs-Nachricht"""
        self._status_label.setText(f"✓ {message}")
        self._status_label.setStyleSheet("color: green; font-weight: bold;")
    
    def _show_error(self, message: str):
        """Zeigt Fehler-Nachricht"""
        self._status_label.setText(f"✗ {message}")
        self._status_label.setStyleSheet("color: red; font-weight: bold;")
    
    def _on_worker_double_clicked(self, row: int, col: int):
        """Handler für Doppelklick auf Worker-Zeile - öffnet Detail-Dialog"""
        if row < 0 or row >= len(self._workers):
            return
        
        worker = self._workers[row]
        
        # Detail-Dialog öffnen
        dialog = WorkerDetailDialog(
            worker=worker,
            analytics_service=self._analytics_service,
            time_entry_repository=self._time_entry_repository,
            capacity_repository=self._capacity_repository,
            parent=self
        )
        dialog.exec()
