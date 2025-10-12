"""
CapacityWidget - UI für Kapazitätsplanung
"""
from typing import Optional, List
from datetime import datetime
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QLabel, QGroupBox, QComboBox, QDateEdit,
    QTextEdit, QProgressBar, QMessageBox, QFileDialog
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor

from ..viewmodels.capacity_viewmodel import CapacityViewModel
from ..models.capacity import Capacity
from ..models.worker import Worker


class CapacityWidget(QWidget):
    """
    Widget für Kapazitätsplanung

    Features:
        - Kapazitäten erstellen/bearbeiten/löschen
        - Worker-basierte Filterung
        - Zeitraum-basierte Ansicht
        - Auslastungsanzeige (Soll/Ist)
    """

    def __init__(self, viewmodel: CapacityViewModel):
        super().__init__()
        self._viewmodel = viewmodel
        self._current_capacity_id: Optional[int] = None
        self._capacities: List[Capacity] = []
        self._workers: List[Worker] = []

        self._setup_ui()
        self._connect_signals()
        self._load_workers()
        self._load_capacities()

    def _setup_ui(self):
        """Erstellt das UI Layout"""
        layout = QHBoxLayout(self)

        # Left: Capacity List & Filter
        left_panel = self._create_list_panel()
        layout.addWidget(left_panel, stretch=2)

        # Right: Capacity Form & Utilization
        right_panel = self._create_form_panel()
        layout.addWidget(right_panel, stretch=1)

    def _create_list_panel(self) -> QWidget:
        """Erstellt das Listen-Panel"""
        panel = QGroupBox("Kapazitäten")
        layout = QVBoxLayout(panel)

        # Filter Section
        filter_layout = QHBoxLayout()

        filter_layout.addWidget(QLabel("Worker:"))
        self._worker_filter = QComboBox()
        self._worker_filter.addItem("Alle Workers", None)
        self._worker_filter.currentIndexChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._worker_filter, stretch=1)

        filter_layout.addWidget(QLabel("Von:"))
        self._start_date_filter = QDateEdit()
        self._start_date_filter.setCalendarPopup(True)
        self._start_date_filter.setDate(QDate.currentDate().addMonths(-3))
        self._start_date_filter.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._start_date_filter)

        filter_layout.addWidget(QLabel("Bis:"))
        self._end_date_filter = QDateEdit()
        self._end_date_filter.setCalendarPopup(True)
        self._end_date_filter.setDate(QDate.currentDate().addMonths(3))
        self._end_date_filter.dateChanged.connect(self._on_filter_changed)
        filter_layout.addWidget(self._end_date_filter)

        layout.addLayout(filter_layout)

        # Capacity Table
        self._capacity_table = QTableWidget()
        self._capacity_table.setColumnCount(6)
        self._capacity_table.setHorizontalHeaderLabels([
            "ID", "Worker", "Von", "Bis", "Stunden", "Auslastung"
        ])
        self._capacity_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._capacity_table.setSelectionBehavior(QTableWidget.SelectRows)
        self._capacity_table.setSelectionMode(QTableWidget.SingleSelection)
        self._capacity_table.itemSelectionChanged.connect(self._on_capacity_selected)
        layout.addWidget(self._capacity_table)

        # Action Buttons
        button_layout = QHBoxLayout()

        self._new_button = QPushButton("Neue Kapazität")
        self._new_button.clicked.connect(self._on_new_clicked)
        button_layout.addWidget(self._new_button)

        self._delete_button = QPushButton("Löschen")
        self._delete_button.clicked.connect(self._on_delete_clicked)
        self._delete_button.setEnabled(False)
        button_layout.addWidget(self._delete_button)
        
        button_layout.addStretch()
        
        # Export Buttons
        self._export_csv_button = QPushButton("📊 Export CSV")
        self._export_csv_button.clicked.connect(self._export_to_csv)
        button_layout.addWidget(self._export_csv_button)
        
        self._export_excel_button = QPushButton("📗 Export Excel")
        self._export_excel_button.clicked.connect(self._export_to_excel)
        button_layout.addWidget(self._export_excel_button)
        
        layout.addLayout(button_layout)

        return panel

    def _create_form_panel(self) -> QWidget:
        """Erstellt das Formular-Panel"""
        panel = QWidget()
        layout = QVBoxLayout(panel)

        # Form Section
        form_group = QGroupBox("Kapazität Details")
        form_layout = QFormLayout(form_group)

        self._worker_input = QComboBox()
        form_layout.addRow("Worker:", self._worker_input)

        self._start_date_input = QDateEdit()
        self._start_date_input.setCalendarPopup(True)
        self._start_date_input.setDate(QDate.currentDate())
        form_layout.addRow("Von:", self._start_date_input)

        self._end_date_input = QDateEdit()
        self._end_date_input.setCalendarPopup(True)
        self._end_date_input.setDate(QDate.currentDate().addMonths(1))
        form_layout.addRow("Bis:", self._end_date_input)

        self._planned_hours_input = QLineEdit()
        self._planned_hours_input.setPlaceholderText("160.0")
        form_layout.addRow("Geplante Stunden:", self._planned_hours_input)

        self._notes_input = QTextEdit()
        self._notes_input.setPlaceholderText("Optionale Notizen...")
        self._notes_input.setMaximumHeight(80)
        form_layout.addRow("Notizen:", self._notes_input)

        layout.addWidget(form_group)

        # Utilization Section
        utilization_group = QGroupBox("Auslastung")
        utilization_layout = QVBoxLayout(utilization_group)

        self._utilization_bar = QProgressBar()
        self._utilization_bar.setTextVisible(True)
        self._utilization_bar.setMinimum(0)
        self._utilization_bar.setMaximum(150)  # 150% als Maximum
        utilization_layout.addWidget(self._utilization_bar)

        stats_layout = QFormLayout()

        self._hours_worked_label = QLabel("-")
        stats_layout.addRow("Gearbeitete Stunden:", self._hours_worked_label)

        self._hours_planned_label = QLabel("-")
        stats_layout.addRow("Geplante Stunden:", self._hours_planned_label)

        self._utilization_label = QLabel("-")
        stats_layout.addRow("Auslastung:", self._utilization_label)

        utilization_layout.addLayout(stats_layout)

        layout.addWidget(utilization_group)

        # Status Label
        self._status_label = QLabel()
        self._status_label.setWordWrap(True)
        layout.addWidget(self._status_label)

        layout.addStretch()

        # Action Buttons
        button_layout = QHBoxLayout()

        self._save_button = QPushButton("Speichern")
        self._save_button.clicked.connect(self._on_save_clicked)
        button_layout.addWidget(self._save_button)

        self._cancel_button = QPushButton("Abbrechen")
        self._cancel_button.clicked.connect(self._on_cancel_clicked)
        button_layout.addWidget(self._cancel_button)

        self._calculate_button = QPushButton("Auslastung berechnen")
        self._calculate_button.clicked.connect(self._on_calculate_clicked)
        button_layout.addWidget(self._calculate_button)

        layout.addLayout(button_layout)

        return panel

    def _connect_signals(self):
        """Verbindet ViewModel Signals"""
        self._viewmodel.capacity_created.connect(self._on_capacity_created)
        self._viewmodel.capacity_updated.connect(self._on_capacity_updated)
        self._viewmodel.capacity_deleted.connect(self._on_capacity_deleted)
        self._viewmodel.capacities_loaded.connect(self._on_capacities_loaded)
        self._viewmodel.utilization_calculated.connect(self._on_utilization_calculated)
        self._viewmodel.validation_failed.connect(self._on_validation_failed)
        self._viewmodel.error_occurred.connect(self._on_error_occurred)

    def _load_workers(self):
        """Lädt Workers für Dropdowns"""
        self._workers = self._viewmodel.get_active_workers()

        # Populate worker dropdowns
        self._worker_input.clear()
        self._worker_filter.clear()
        self._worker_filter.addItem("Alle Workers", None)

        for worker in self._workers:
            self._worker_input.addItem(worker.name, worker.id)
            self._worker_filter.addItem(worker.name, worker.id)

    def _load_capacities(self):
        """Lädt Kapazitäten basierend auf Filter"""
        worker_id = self._worker_filter.currentData()
        start_date = self._start_date_filter.date().toPython()
        end_date = self._end_date_filter.date().toPython()

        start_datetime = datetime(start_date.year, start_date.month, start_date.day)
        end_datetime = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)

        if worker_id:
            self._viewmodel.load_capacities_for_worker(
                worker_id, start_datetime, end_datetime
            )
        else:
            self._viewmodel.load_all_capacities(start_datetime, end_datetime)

    def _populate_table(self, capacities: List[Capacity]):
        """Füllt die Tabelle mit Kapazitäten"""
        self._capacity_table.setRowCount(0)
        self._capacities = capacities

        for capacity in capacities:
            row = self._capacity_table.rowCount()
            self._capacity_table.insertRow(row)

            # Find worker name
            worker = next(
                (w for w in self._workers if w.id == capacity.worker_id),
                None
            )
            worker_name = worker.name if worker else f"Worker #{capacity.worker_id}"

            # ID
            id_item = QTableWidgetItem(str(capacity.id))
            id_item.setData(Qt.UserRole, capacity.id)
            id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)
            self._capacity_table.setItem(row, 0, id_item)

            # Worker
            worker_item = QTableWidgetItem(worker_name)
            worker_item.setFlags(worker_item.flags() & ~Qt.ItemIsEditable)
            self._capacity_table.setItem(row, 1, worker_item)

            # Start Date
            start_str = capacity.start_date.strftime("%d.%m.%Y")
            start_item = QTableWidgetItem(start_str)
            start_item.setFlags(start_item.flags() & ~Qt.ItemIsEditable)
            self._capacity_table.setItem(row, 2, start_item)

            # End Date
            end_str = capacity.end_date.strftime("%d.%m.%Y")
            end_item = QTableWidgetItem(end_str)
            end_item.setFlags(end_item.flags() & ~Qt.ItemIsEditable)
            self._capacity_table.setItem(row, 3, end_item)

            # Planned Hours
            hours_item = QTableWidgetItem(f"{capacity.planned_hours:.1f}h")
            hours_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            hours_item.setFlags(hours_item.flags() & ~Qt.ItemIsEditable)
            self._capacity_table.setItem(row, 4, hours_item)

            # Calculate and display utilization
            utilization = self._calculate_capacity_utilization(capacity)
            util_item = QTableWidgetItem(utilization['display'])
            util_item.setTextAlignment(Qt.AlignCenter)
            util_item.setFlags(util_item.flags() & ~Qt.ItemIsEditable)

            # Color coding based on utilization
            if utilization['percent'] is not None:
                if utilization['percent'] < 80:
                    util_item.setForeground(QColor("orange"))
                elif utilization['percent'] <= 110:
                    util_item.setForeground(QColor("green"))
                else:
                    util_item.setForeground(QColor("red"))

            self._capacity_table.setItem(row, 5, util_item)

    def _clear_form(self):
        """Leert das Formular"""
        self._current_capacity_id = None
        self._worker_input.setCurrentIndex(0)
        self._start_date_input.setDate(QDate.currentDate())
        self._end_date_input.setDate(QDate.currentDate().addMonths(1))
        self._planned_hours_input.clear()
        self._notes_input.clear()
        self._clear_utilization()
        self._status_label.clear()

    def _populate_form(self, capacity: Capacity):
        """Füllt das Formular mit Kapazitäts-Daten"""
        self._current_capacity_id = capacity.id

        # Set worker
        index = self._worker_input.findData(capacity.worker_id)
        if index >= 0:
            self._worker_input.setCurrentIndex(index)

        # Set dates
        start_qdate = QDate(
            capacity.start_date.year,
            capacity.start_date.month,
            capacity.start_date.day
        )
        self._start_date_input.setDate(start_qdate)

        end_qdate = QDate(
            capacity.end_date.year,
            capacity.end_date.month,
            capacity.end_date.day
        )
        self._end_date_input.setDate(end_qdate)

        # Set hours
        self._planned_hours_input.setText(str(capacity.planned_hours))

        # Set notes
        self._notes_input.setPlainText(capacity.notes or "")

        self._status_label.clear()

        # Auto-calculate utilization
        self._calculate_utilization()

    def _clear_utilization(self):
        """Leert die Auslastungs-Anzeige"""
        self._utilization_bar.setValue(0)
        self._utilization_bar.setFormat("-")
        self._hours_worked_label.setText("-")
        self._hours_planned_label.setText("-")
        self._utilization_label.setText("-")

    def _calculate_capacity_utilization(self, capacity: Capacity) -> dict:
        """
        Berechnet die Auslastung für eine einzelne Kapazität

        Args:
            capacity: Die Kapazität für die die Auslastung berechnet werden soll

        Returns:
            Dict mit 'percent' (float oder None) und 'display' (str)
        """
        try:
            # Direkter Aufruf des AnalyticsService ohne Signal-Emission
            # um Probleme mit dem Form-Panel zu vermeiden
            analytics_service = self._viewmodel._analytics_service
            utilization_data = analytics_service.calculate_worker_utilization(
                capacity.worker_id,
                capacity.start_date,
                capacity.end_date
            )

            if utilization_data and utilization_data.get('hours_planned', 0) > 0:
                percent = utilization_data['utilization_percent']
                return {
                    'percent': percent,
                    'display': f"{percent:.1f}%"
                }
            else:
                # Keine Daten oder keine geplanten Stunden
                return {
                    'percent': None,
                    'display': "-"
                }
        except Exception:
            # Bei Fehler einfach "-" anzeigen
            return {
                'percent': None,
                'display': "-"
            }

    def _calculate_utilization(self):
        """Berechnet die Auslastung für aktuellen Form-Zustand"""
        worker_id = self._worker_input.currentData()
        if not worker_id:
            return

        start_date = self._start_date_input.date().toPython()
        end_date = self._end_date_input.date().toPython()

        start_datetime = datetime(start_date.year, start_date.month, start_date.day)
        end_datetime = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)

        self._viewmodel.calculate_utilization(worker_id, start_datetime, end_datetime)

    def _on_filter_changed(self):
        """Handler für Filter-Änderungen"""
        self._load_capacities()

    def _on_capacity_selected(self):
        """Handler für Kapazitäts-Auswahl"""
        selected_rows = self._capacity_table.selectedItems()
        if not selected_rows:
            self._delete_button.setEnabled(False)
            return

        self._delete_button.setEnabled(True)

        row = selected_rows[0].row()
        capacity_id = self._capacity_table.item(row, 0).data(Qt.UserRole)

        capacity = next((c for c in self._capacities if c.id == capacity_id), None)
        if capacity:
            self._populate_form(capacity)

    def _on_new_clicked(self):
        """Handler für Neu-Button"""
        self._clear_form()
        self._capacity_table.clearSelection()

    def _on_save_clicked(self):
        """Handler für Speichern-Button"""
        worker_id = self._worker_input.currentData()
        if not worker_id:
            self._show_error("Bitte wählen Sie einen Worker")
            return

        start_date = self._start_date_input.date().toPython()
        end_date = self._end_date_input.date().toPython()
        start_datetime = datetime(start_date.year, start_date.month, start_date.day)
        end_datetime = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)

        try:
            planned_hours = float(self._planned_hours_input.text())
        except ValueError:
            self._show_error("Geplante Stunden müssen eine Zahl sein")
            return

        notes = self._notes_input.toPlainText()

        if self._current_capacity_id is None:
            # Create new
            self._viewmodel.create_capacity(
                worker_id, start_datetime, end_datetime, planned_hours, notes
            )
        else:
            # Update existing
            self._viewmodel.update_capacity(
                self._current_capacity_id, worker_id,
                start_datetime, end_datetime, planned_hours, notes
            )

    def _on_cancel_clicked(self):
        """Handler für Abbrechen-Button"""
        self._clear_form()
        self._capacity_table.clearSelection()

    def _on_calculate_clicked(self):
        """Handler für Auslastung-Berechnen-Button"""
        self._calculate_utilization()

    def _on_delete_clicked(self):
        """Handler für Löschen-Button"""
        if self._current_capacity_id is None:
            return

        reply = QMessageBox.question(
            self,
            "Kapazität löschen",
            "Möchten Sie diese Kapazität wirklich löschen?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self._viewmodel.delete_capacity(self._current_capacity_id)

    def _on_capacity_created(self, capacity_id: int):
        """Handler für capacity_created Signal"""
        self._show_success(f"Kapazität erfolgreich erstellt (ID: {capacity_id})")
        self._clear_form()
        self._load_capacities()

    def _on_capacity_updated(self, capacity_id: int):
        """Handler für capacity_updated Signal"""
        self._show_success("Kapazität erfolgreich aktualisiert")
        self._load_capacities()

    def _on_capacity_deleted(self, capacity_id: int):
        """Handler für capacity_deleted Signal"""
        self._show_success("Kapazität erfolgreich gelöscht")
        self._clear_form()
        self._load_capacities()

    def _on_capacities_loaded(self, capacities: List[Capacity]):
        """Handler für capacities_loaded Signal"""
        self._populate_table(capacities)

    def _on_utilization_calculated(
        self, worker_id: int, utilization: float, hours_worked: float, hours_planned: float
    ):
        """Handler für utilization_calculated Signal"""
        # Update progress bar
        self._utilization_bar.setValue(int(utilization))
        self._utilization_bar.setFormat(f"{utilization:.1f}%")

        # Color coding
        if utilization < 80:
            self._utilization_bar.setStyleSheet("QProgressBar::chunk { background-color: orange; }")
        elif utilization <= 110:
            self._utilization_bar.setStyleSheet("QProgressBar::chunk { background-color: green; }")
        else:
            self._utilization_bar.setStyleSheet("QProgressBar::chunk { background-color: red; }")

        # Update labels
        self._hours_worked_label.setText(f"{hours_worked:.1f}h")
        self._hours_planned_label.setText(f"{hours_planned:.1f}h")
        self._utilization_label.setText(f"{utilization:.1f}%")

    def _on_validation_failed(self, message: str):
        """Handler für validation_failed Signal"""
        self._show_error(message)

    def _on_error_occurred(self, message: str):
        """Handler für error_occurred Signal"""
        self._show_error(message)

    def _show_success(self, message: str):
        """Zeigt Erfolgs-Nachricht"""
        self._status_label.setText(f"✓ {message}")
        self._status_label.setStyleSheet("color: green;")

    def _show_error(self, message: str):
        """Zeigt Fehler-Nachricht"""
        self._status_label.setText(f"✗ {message}")
        self._status_label.setStyleSheet("color: red;")
    
    def _export_to_csv(self):
        """Exportiert Kapazitätsdaten als CSV"""
        if not self._capacities:
            QMessageBox.warning(
                self,
                "Keine Daten",
                "Es sind keine Kapazitätsdaten zum Exportieren vorhanden."
            )
            return
        
        # Get filter information for filename and header
        start_date = self._start_date_filter.date().toPython()
        end_date = self._end_date_filter.date().toPython()
        worker_filter = self._worker_filter.currentText()
        
        # Datei-Dialog
        default_filename = f"capacity_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export als CSV",
            default_filename,
            "CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                
                # Bericht-Informationen
                writer.writerow(['Kapazitätsplanung Bericht'])
                writer.writerow(['Zeitraum', f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"])
                writer.writerow(['Worker-Filter', worker_filter])
                writer.writerow(['Export-Datum', datetime.now().strftime('%d.%m.%Y %H:%M:%S')])
                writer.writerow([])
                
                # Header
                writer.writerow([
                    'ID', 'Worker', 'Von', 'Bis', 'Geplante Stunden',
                    'Tage', 'Stunden/Tag', 'Notizen'
                ])
                
                # Daten
                total_hours = 0.0
                for capacity in self._capacities:
                    # Find worker name
                    worker = next((w for w in self._workers if w.id == capacity.worker_id), None)
                    worker_name = worker.name if worker else f"Worker #{capacity.worker_id}"
                    
                    writer.writerow([
                        capacity.id,
                        worker_name,
                        capacity.start_date.strftime('%d.%m.%Y'),
                        capacity.end_date.strftime('%d.%m.%Y'),
                        f"{capacity.planned_hours:.1f}",
                        capacity.days_count(),
                        f"{capacity.hours_per_day():.1f}",
                        capacity.notes or "-"
                    ])
                    total_hours += capacity.planned_hours
                
                # Zusammenfassung
                writer.writerow([])
                writer.writerow(['Zusammenfassung'])
                writer.writerow(['Anzahl Einträge', len(self._capacities)])
                writer.writerow(['Gesamt geplante Stunden', f"{total_hours:.1f}"])
            
            QMessageBox.information(
                self,
                "Export erfolgreich",
                f"Kapazitätsdaten wurden erfolgreich exportiert:\n\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export fehlgeschlagen",
                f"Fehler beim Exportieren der Daten:\n\n{str(e)}"
            )
    
    def _export_to_excel(self):
        """Exportiert Kapazitätsdaten als Excel mit Formatierung"""
        if not self._capacities:
            QMessageBox.warning(
                self,
                "Keine Daten",
                "Es sind keine Kapazitätsdaten zum Exportieren vorhanden."
            )
            return
        
        # Get filter information for filename and header
        start_date = self._start_date_filter.date().toPython()
        end_date = self._end_date_filter.date().toPython()
        worker_filter = self._worker_filter.currentText()
        
        # Datei-Dialog
        default_filename = f"capacity_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export als Excel",
            default_filename,
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Kapazitätsplanung"
            
            # Titel und Berichtsinformationen
            title_cell = ws.cell(row=1, column=1, value="Kapazitätsplanung Bericht")
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws.merge_cells('A1:H1')
            
            ws.cell(row=2, column=1, value="Zeitraum:")
            ws.cell(row=2, column=2, value=f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
            ws.cell(row=3, column=1, value="Worker-Filter:")
            ws.cell(row=3, column=2, value=worker_filter)
            ws.cell(row=4, column=1, value="Export-Datum:")
            ws.cell(row=4, column=2, value=datetime.now().strftime('%d.%m.%Y %H:%M:%S'))
            
            # Header-Zeile (Row 6)
            headers = ['ID', 'Worker', 'Von', 'Bis', 'Geplante Stunden', 
                      'Tage', 'Stunden/Tag', 'Notizen']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col_num, value=header)
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Daten-Zeilen
            row_num = 7
            total_hours = 0.0
            
            for capacity in self._capacities:
                # Find worker name
                worker = next((w for w in self._workers if w.id == capacity.worker_id), None)
                worker_name = worker.name if worker else f"Worker #{capacity.worker_id}"
                
                row_data = [
                    capacity.id,
                    worker_name,
                    capacity.start_date.strftime('%d.%m.%Y'),
                    capacity.end_date.strftime('%d.%m.%Y'),
                    capacity.planned_hours,
                    capacity.days_count(),
                    capacity.hours_per_day(),
                    capacity.notes or "-"
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    
                    # Alignment
                    if col_num in [1, 5, 6, 7]:  # ID, Hours, Days, Hours/Day - center
                        cell.alignment = Alignment(horizontal='center')
                    elif col_num in [3, 4]:  # Dates - center
                        cell.alignment = Alignment(horizontal='center')
                    else:  # Worker, Notes - left
                        cell.alignment = Alignment(horizontal='left')
                    
                    # Border
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Number formatting for hours
                    if col_num in [5, 7]:  # Planned Hours, Hours/Day
                        cell.number_format = '0.0'
                
                total_hours += capacity.planned_hours
                row_num += 1
            
            # Zusammenfassung (mit Abstand)
            row_num += 2
            summary_row = row_num
            
            summary_cell = ws.cell(row=summary_row, column=1, value="Zusammenfassung")
            summary_cell.font = Font(bold=True, size=14, color="FFFFFF")
            summary_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
            
            summary_data = [
                ('Anzahl Einträge', len(self._capacities)),
                ('Gesamt geplante Stunden', f"{total_hours:.1f}")
            ]
            
            for i, (label, value) in enumerate(summary_data, 1):
                row = summary_row + i
                label_cell = ws.cell(row=row, column=1, value=label)
                label_cell.font = Font(bold=True)
                
                value_cell = ws.cell(row=row, column=2, value=value)
                value_cell.alignment = Alignment(horizontal='left')
            
            # Spaltenbreiten anpassen
            ws.column_dimensions['A'].width = 8   # ID
            ws.column_dimensions['B'].width = 20  # Worker
            ws.column_dimensions['C'].width = 12  # Von
            ws.column_dimensions['D'].width = 12  # Bis
            ws.column_dimensions['E'].width = 16  # Geplante Stunden
            ws.column_dimensions['F'].width = 8   # Tage
            ws.column_dimensions['G'].width = 14  # Stunden/Tag
            ws.column_dimensions['H'].width = 30  # Notizen
            
            # Speichern
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Export erfolgreich",
                f"Kapazitätsdaten wurden erfolgreich exportiert:\n\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export fehlgeschlagen",
                f"Fehler beim Exportieren der Daten:\n\n{str(e)}"
            )
